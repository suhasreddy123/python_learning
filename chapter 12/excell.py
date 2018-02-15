import openpyxl
wb = openpyxl.load_workbook('CONCRETE QTY FOR ABUTMENTS.xlsx')
print(wb.get_sheet_names())
sheet=wb.get_sheet_by_name('5.295')
print (sheet)
print(sheet['C5'])
print(sheet['C5'].value)
print(sheet.cell(row=1,column=2))
#for i in range(1,20,1):
    #print(i,sheet.cell(row=i,column=5).value)
print(sheet.max_row)
print(sheet.max_column)